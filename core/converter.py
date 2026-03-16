import os
from openpyxl import load_workbook
from lxml import etree
from core.validator import *

class XliffConverter:
    def __init__(self):
        pass

    def export_excel(self, source_excel, target_excel, output_xliff, source_lang, target_lang, hidden, progress_callback=None, blank_if_equal=True):
        wb_source = load_workbook(source_excel, data_only=False)
        wb_target = load_workbook(target_excel, data_only=False)

        root = etree.Element("xliff", version="1.2")
        file_elem = etree.SubElement(root, "file", attrib={
            "source-language": source_lang,
            "target-language": target_lang,
            "datatype": "plaintext",
            "original": os.path.basename(source_excel)
        })
        body = etree.SubElement(file_elem, "body")

        total_cells = sum(
            1
            for ws in wb_source.worksheets
            if ws.sheet_state == "visible"
            for row in ws.iter_rows()
            if not ws.row_dimensions[row[0].row].hidden
            for cell in row
            if not ws.column_dimensions[cell.column_letter].hidden
        )

        counter = 0
        unit_id = 1

        for sheet_name in wb_source.sheetnames:

            ws_source = wb_source[sheet_name]
            ws_target = wb_target[sheet_name]

            # Skip hidden sheets
            if hidden and ws_source.sheet_state != "visible":
                continue

            for row in ws_source.iter_rows():
                row_index = row[0].row

                if hidden and ws_source.row_dimensions[row_index].hidden:
                    continue

                for cell in row:

                    col_letter = cell.column_letter

                    if hidden and ws_source.column_dimensions[col_letter].hidden:
                        continue

                    src_value = cell.value
                    tgt_value = ws_target[cell.coordinate].value

                    src = "" if src_value is None else str(src_value)
                    tgt = "" if tgt_value is None else str(tgt_value)

                    c_hash = generate_hash(src) 

                    is_formula = cell.data_type == "f"

                    if blank_if_equal and src.strip() == tgt.strip():
                        tgt = ""

                    trans_unit = etree.SubElement(
                        body,
                        "trans-unit",
                        attrib={
                            "id": str(unit_id),
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "hash": c_hash,
                            "formula": str(is_formula)
                        }
                    )

                    etree.SubElement(trans_unit, "source").text = src
                    etree.SubElement(trans_unit, "target").text = tgt

                    trans_unit.set(
                        "state",
                        "translated" if tgt.strip() else "needs-translation"
                    )

                    unit_id += 1
                    counter += 1

                    if progress_callback:
                        progress_callback(counter, total_cells)

        tree = etree.ElementTree(root)
        tree.write(output_xliff, pretty_print=True, encoding="utf-8", xml_declaration=True)

    def import_xliff(self, original_excel, xliff_file, output_excel, progress_callback=None):
        
        # 🔹 PHASE 1 — VALIDATION
        if progress_callback:
            progress_callback(0, 1)

        '''is_valid = validate_structure(
            original_excel,
            xliff_file,
            verbose=False,
            progress_callback=progress_callback
        )

        if not is_valid:
            raise Exception("Import aborted due to validation failure.")
        '''

        # 🔹 PHASE 2 — IMPORT
        
        wb = load_workbook(original_excel, data_only=False)
        tree = etree.parse(xliff_file)
        root = tree.getroot()

        trans_units = root.xpath(".//trans-unit")
        total = len(trans_units)

        is_valid, errors = validate(original_excel, trans_units)
        if not is_valid:
            return False, errors
        
        for i, trans_unit in enumerate(trans_units):

            sheet = trans_unit.get("sheet")
            cell_ref = trans_unit.get("cell")
            is_formula = trans_unit.get("formula") == "True"
            target = trans_unit.findtext("target")

            if target and not is_formula:
                ws = wb[sheet]
                ws[cell_ref].value = target

            if progress_callback:
                progress_callback(i + 1, total)

        wb.save(output_excel)
        return True, []