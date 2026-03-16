import hashlib
from openpyxl import load_workbook

def generate_hash(text):
    return hashlib.md5(str(text or "").encode('utf-8')).hexdigest()

def validate(original_excel, trans_units):
    wb = load_workbook(original_excel, data_only=True, read_only=True)
    errors = []

    for tu in trans_units:
        sheet_name = tu.get("sheet")
        cell_ref = tu.get("cell")
        expected_hash = tu.get("hash")
        
        if sheet_name not in wb.sheetnames:
            errors.append(f"Hoja faltante: {sheet_name}")
            continue
            
        current_val = wb[sheet_name][cell_ref].value
        if generate_hash(current_val) != expected_hash:
            errors.append(f"Source mismatch: {sheet_name}!{cell_ref} (Content modified)")

    return len(errors) == 0, errors